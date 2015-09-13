from lxml import etree
from openpyxl import load_workbook
ur=[]
u1=[]
u2=[]
u3=[]
wb=load_workbook('/users/yufangxian/Documents/exercise/mm1.xlsx')

sheet_names=wb.get_sheet_names()
ws=wb.get_sheet_by_name(sheet_names[0])

for irow in range(1,4):
    for icol in range(1,4):
        ur.append(ws.cell(row=irow,column=icol).value)

u1=ur[0:3]
u2=ur[3:6]
u3=ur[6:9]
print(u1)
print(u2)
print(u3)
print(ur)
ur_4=(str(u1)+','+str(u2)+','+str(u3))
print(ur_4)
root=etree.Element('root')
numbers=etree.SubElement(root,"numbers").text

root.append(etree.Comment("some comment"))
etree.SubElement(root,"numbers").text=ur_4

print(etree.tostring(root,pretty_print=True))
with open('/users/yufangxian/Documents/exercise/xxx.xml','wb') as f:
    f.write(etree.tostring(root,pretty_print=True))
